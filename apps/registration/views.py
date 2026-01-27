"""
Views untuk Registration (PKBM - Final Version).
FLOW: Create → Documents → Review → Submit → Payment
"""
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import DetailView, ListView
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponse
from django.views.decorators.http import require_POST

from .models import StudentRegistration, Document
from .forms import StudentRegistrationForm, DocumentUploadForm
from .services import RegistrationService
from apps.accounts.permissions import StaffRequiredMixin

import logging

logger = logging.getLogger('apps.registration')


# ============================================
# PUBLIC VIEWS (Tidak perlu login)
# ============================================

class CreateRegistrationView(View):
    """STEP 1: Buat pendaftaran (biodata)"""
    
    template_name = 'registration/create.html'
    
    def get(self, request):
        academic_year = self._get_current_academic_year()
        form = StudentRegistrationForm()
        
        return render(request, self.template_name, {
            'form': form,
            'academic_year': academic_year,
        })
    
    def post(self, request):
        academic_year = self._get_current_academic_year()
        form = StudentRegistrationForm(request.POST)
        
        if form.is_valid():
            try:
                # Buat registration object
                registration = StudentRegistration(
                    academic_year=academic_year,
                    status=StudentRegistration.RegistrationStatus.DRAFT,
                    **form.cleaned_data
                )
                
                # SAVE DULU (ini yang generate registration_number)
                registration.save()
                
                # SEKARANG registration_number sudah ada, BARU show message
                messages.success(
                    request,
                    f'Pendaftaran berhasil dibuat! {registration.registration_number}. '
                    f'Silakan upload dokumen yang diperlukan untuk mendapatkan nomor pendaftaran.'
                )
                
                logger.info(f"Registration created: {registration.registration_number}")
                
                # Redirect ke documents
                return redirect('registration:documents', pk=registration.id)
                
            except Exception as e:
                logger.error(f"Failed to create registration", exc_info=True)
                messages.error(request, f'Gagal membuat pendaftaran: {str(e)}')
        else:
            messages.error(request, 'Terjadi kesalahan. Periksa kembali form Anda.')
        
        return render(request, self.template_name, {
            'form': form,
            'academic_year': academic_year,
        })
    
    def _get_current_academic_year(self):
        from datetime import datetime
        now = datetime.now()
        if now.month >= 7:
            return f"{now.year}/{now.year + 1}"
        else:
            return f"{now.year - 1}/{now.year}"

class DocumentUploadView(View):
    """STEP 2: Upload dokumen"""
    
    template_name = 'registration/documents.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.registration = get_object_or_404(StudentRegistration, pk=kwargs.get('pk'))
        
        if self.registration.status != StudentRegistration.RegistrationStatus.DRAFT:
            messages.error(request, 'Dokumen tidak bisa diupload. Pendaftaran sudah disubmit.')
            return redirect('registration:check_status')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        form = DocumentUploadForm(registration=self.registration)
        documents = self.registration.documents.all()
        
        required_docs = ['KTP', 'KK', 'AKTA']
        uploaded_types = set(documents.values_list('document_type', flat=True))
        missing_docs = set(required_docs) - uploaded_types
        
        missing_doc_labels = []
        for doc_type in missing_docs:
            try:
                label = Document.DocumentType(doc_type).label
            except ValueError:
                label = doc_type
            missing_doc_labels.append(label)
        
        can_proceed = not missing_docs
        
        return render(request, self.template_name, {
            'form': form,
            'registration': self.registration,
            'documents': documents,
            'missing_documents': missing_doc_labels,
            'can_proceed': can_proceed,
        })
    
    def post(self, request, *args, **kwargs):
        form = DocumentUploadForm(request.POST, request.FILES, registration=self.registration)
        
        if form.is_valid():
            try:
                document = form.save()
                messages.success(request, f'Dokumen {document.get_document_type_display()} berhasil diupload.')
                logger.info(f"Document uploaded: {document.document_type} for {self.registration.registration_number}")
                return redirect('registration:documents', pk=self.registration.id)
            except Exception as e:
                logger.error(f"Failed to upload document", exc_info=True)
                messages.error(request, 'Gagal mengupload dokumen. Silakan coba lagi.')
        else:
            messages.error(request, 'File tidak valid. Periksa kembali file Anda.')
        
        documents = self.registration.documents.all()
        return render(request, self.template_name, {
            'form': form,
            'registration': self.registration,
            'documents': documents,
        })


class ReviewRegistrationView(View):
    """STEP 3: Review data + Checkbox"""
    
    template_name = 'registration/review.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.registration = get_object_or_404(StudentRegistration, pk=kwargs.get('pk'))
        
        if self.registration.status != StudentRegistration.RegistrationStatus.DRAFT:
            messages.error(request, 'Pendaftaran sudah disubmit.')
            return redirect('registration:check_status')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        documents = self.registration.documents.all()
        
        required_docs = ['KTP', 'KK', 'AKTA']
        uploaded_types = set(documents.values_list('document_type', flat=True))
        missing_docs = set(required_docs) - uploaded_types
        
        if missing_docs:
            messages.error(request, 'Dokumen wajib belum lengkap. Silakan upload terlebih dahulu.')
            return redirect('registration:documents', pk=self.registration.id)
        
        return render(request, self.template_name, {
            'registration': self.registration,
            'documents': documents,
        })


@require_POST
def submit_registration_view(request, pk):
    """Submit pendaftaran"""
    
    registration = get_object_or_404(StudentRegistration, pk=pk)
    
    if registration.status != StudentRegistration.RegistrationStatus.DRAFT:
        messages.error(request, 'Pendaftaran sudah disubmit sebelumnya.')
        return redirect('registration:check_status')
    
    declaration_confirmed = request.POST.get('declaration_confirmed')
    
    if not declaration_confirmed:
        messages.error(request, 'Anda harus menyetujui pernyataan untuk melanjutkan.')
        return redirect('registration:review', pk=registration.id)
    
    try:
        registration.declaration_agreed = True
        registration.declaration_agreed_at = timezone.now()
        registration.save()
    except AttributeError:
        pass
    
    try:
        # DEBUG LOG
        logger.info(f"BEFORE submit - RegID: {registration.id}, Number: '{registration.registration_number}'")
        
        # Submit (generate nomor)
        RegistrationService.submit_registration(registration)
        
        # Refresh
        registration.refresh_from_db()
        
        # DEBUG LOG
        logger.info(f"AFTER submit - Number: '{registration.registration_number}', Status: {registration.status}")
        
        # Verify
        if not registration.registration_number:
            raise ValueError('Nomor pendaftaran gagal di-generate.')
        
        # Success message
        messages.success(
            request,
            f'Pendaftaran berhasil! Nomor: {registration.registration_number}. '
            f'Lanjutkan ke pembayaran.'
        )
        
        logger.info(f"SUCCESS - Redirecting to payment for: {registration.id}")
        
        # REDIRECT KE PAYMENT
        redirect_url = f'/payments/create/{registration.id}/'
        logger.info(f"Redirect URL: {redirect_url}")
        
        return redirect('payments:create', registration_id=registration.id)
        
    except ValueError as e:
        logger.error(f"ValueError: {str(e)}")
        messages.error(request, str(e))
        return redirect('registration:review', pk=registration.id)
        
    except Exception as e:
        logger.error(f"UNEXPECTED ERROR: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('registration:review', pk=registration.id)

def check_status_view(request):
    """Cek status pendaftaran - GET & POST"""
    
    if request.method == 'GET':
        # Tampilkan form
        return render(request, 'registration/check_status.html')
    
    elif request.method == 'POST':
        # Process form
        registration_number = request.POST.get('registration_number', '').strip()
        identifier = request.POST.get('identifier', '').strip()
        
        if not registration_number or not identifier:
            messages.error(request, 'Mohon isi semua field.')
            return render(request, 'registration/check_status.html')
        
        registration = StudentRegistration.objects.filter(
            registration_number=registration_number
        ).filter(
            Q(nik=identifier) |
            Q(nisn=identifier) |
            Q(contact_email__iexact=identifier) |
            Q(contact_phone=identifier) |
            Q(parent_phone=identifier)
        ).first()
        
        if not registration:
            messages.error(
                request,
                'Data tidak ditemukan. Periksa kembali nomor pendaftaran dan identitas Anda.'
            )
            return render(request, 'registration/check_status.html')
        
        # Get payment if exists
        payment = None
        try:
            payment = registration.payment
        except:
            pass
        
        return render(request, 'registration/status_result.html', {
            'registration': registration,
            'payment': payment,
        })


# ============================================
# STAFF VIEWS - Dashboard & Verification
# ============================================

class StaffDashboardView(LoginRequiredMixin, StaffRequiredMixin, View):
    """STAFF ONLY - Dashboard"""
    
    template_name = 'registration/staff/dashboard.html'
    
    def get(self, request):
        stats = {
            'total': StudentRegistration.objects.count(),
            'draft': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.DRAFT).count(),
            'submitted': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.SUBMITTED).count(),
            'expired': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.PAYMENT_EXPIRED).count(),
            'paid': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.PAID).count(),
            'verified': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.VERIFIED).count(),
            'rejected': StudentRegistration.objects.filter(status=StudentRegistration.RegistrationStatus.REJECTED).count(),
        }
        
        recent_paid = StudentRegistration.objects.filter(
            status=StudentRegistration.RegistrationStatus.PAID
        ).order_by('-submitted_at')[:10]
        
        recent_verified = StudentRegistration.objects.filter(
            status=StudentRegistration.RegistrationStatus.VERIFIED
        ).order_by('-verified_at')[:5]
        
        program_stats = StudentRegistration.objects.values('program_choice').annotate(count=Count('id')).order_by('-count')
        
        return render(request, self.template_name, {
            'stats': stats,
            'recent_paid': recent_paid,
            'recent_verified': recent_verified,
            'program_stats': program_stats,
        })


class RegistrationListView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    """STAFF ONLY - List pendaftaran"""
    
    model = StudentRegistration
    template_name = 'registration/staff/list.html'
    context_object_name = 'registrations'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = StudentRegistration.objects.select_related('verified_by').prefetch_related('documents').order_by('-created_at')
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        program = self.request.GET.get('program')
        if program:
            queryset = queryset.filter(program_choice=program)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(registration_number__icontains=search) |
                Q(full_name__icontains=search) |
                Q(nik__icontains=search) |
                Q(nisn__icontains=search) |
                Q(contact_email__icontains=search) |
                Q(contact_phone__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = StudentRegistration.RegistrationStatus.choices
        context['program_choices'] = StudentRegistration.ProgramChoice.choices
        context['current_status'] = self.request.GET.get('status', '')
        context['current_program'] = self.request.GET.get('program', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class StaffRegistrationDetailView(LoginRequiredMixin, StaffRequiredMixin, DetailView):
    """STAFF ONLY - Detail pendaftaran"""
    
    model = StudentRegistration
    template_name = 'registration/staff/detail.html'
    context_object_name = 'registration'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        registration = self.object
        context['documents'] = registration.documents.all()
        
        required_docs = ['KTP', 'KK', 'AKTA']
        uploaded_types = set(registration.documents.values_list('document_type', flat=True))
        missing_docs = set(required_docs) - uploaded_types
        
        context['missing_documents'] = [Document.DocumentType(doc).label for doc in missing_docs]
        context['documents_complete'] = not missing_docs
        
        if hasattr(registration, 'payment'):
            context['payment'] = registration.payment
        
        return context


class VerifyRegistrationView(LoginRequiredMixin, StaffRequiredMixin, View):
    """STAFF ONLY - Verify/Reject"""
    
    def post(self, request, pk):
        registration = get_object_or_404(StudentRegistration, pk=pk)
        
        # Hanya PAID yang bisa diverifikasi
        if registration.status != StudentRegistration.RegistrationStatus.PAID:
            messages.error(
                request,
                'Hanya pendaftaran yang sudah dibayar yang bisa diverifikasi.'
            )
            return redirect('registration:staff_detail', pk=registration.id)
        
        action = request.POST.get('action')
        notes = request.POST.get('verification_notes', '').strip()
        
        try:
            if action == 'approve':
                registration.status = StudentRegistration.RegistrationStatus.VERIFIED
                registration.verified_at = timezone.now()
                registration.verified_by = request.user
                registration.verification_notes = notes or 'Pendaftaran disetujui'
                registration.save()
                
                messages.success(request, f'Pendaftaran {registration.registration_number} DISETUJUI.')
                logger.info(f"Registration APPROVED: {registration.registration_number} by {request.user}")
                
            elif action == 'reject':
                if not notes:
                    messages.error(request, 'Alasan penolakan wajib diisi.')
                    return redirect('registration:staff_detail', pk=registration.id)
                
                registration.status = StudentRegistration.RegistrationStatus.REJECTED
                registration.verified_at = timezone.now()
                registration.verified_by = request.user
                registration.verification_notes = notes
                registration.save()
                
                messages.warning(request, f'Pendaftaran {registration.registration_number} DITOLAK.')
                logger.info(f"Registration REJECTED: {registration.registration_number} by {request.user}")
            
            else:
                messages.error(request, 'Aksi tidak valid.')
            
        except Exception as e:
            logger.error(f"Verification error: {str(e)}", exc_info=True)
            messages.error(request, f'Gagal memverifikasi: {str(e)}')
        
        return redirect('registration:staff_detail', pk=registration.id)


class BulkVerifyView(LoginRequiredMixin, StaffRequiredMixin, View):
    """STAFF ONLY - Bulk approve/reject"""
    
    def post(self, request):
        action = request.POST.get('action')
        registration_ids = request.POST.getlist('registration_ids')
        
        if not registration_ids:
            messages.error(request, 'Pilih minimal 1 pendaftaran.')
            return redirect('registration:staff_list')
        
        try:
            registrations = StudentRegistration.objects.filter(
                id__in=registration_ids,
                status=StudentRegistration.RegistrationStatus.PAID
            )
            
            count = registrations.count()
            
            if action == 'bulk_approve':
                registrations.update(
                    status=StudentRegistration.RegistrationStatus.VERIFIED,
                    verified_at=timezone.now(),
                    verified_by=request.user,
                    verification_notes='Bulk approval'
                )
                
                messages.success(request, f'{count} pendaftaran berhasil disetujui.')
                logger.info(f"Bulk approved {count} registrations by {request.user}")
                
            elif action == 'bulk_reject':
                notes = request.POST.get('bulk_notes', '').strip()
                if not notes:
                    messages.error(request, 'Alasan penolakan wajib diisi untuk bulk reject.')
                    return redirect('registration:staff_list')
                
                registrations.update(
                    status=StudentRegistration.RegistrationStatus.REJECTED,
                    verified_at=timezone.now(),
                    verified_by=request.user,
                    verification_notes=notes
                )
                
                messages.warning(request, f'{count} pendaftaran berhasil ditolak.')
                logger.info(f"Bulk rejected {count} registrations by {request.user}")
            
        except Exception as e:
            logger.error(f"Bulk verification error: {str(e)}", exc_info=True)
            messages.error(request, f'Gagal: {str(e)}')
        
        return redirect('registration:staff_list')


class ExportRegistrationsView(LoginRequiredMixin, StaffRequiredMixin, View):
    """STAFF ONLY - Export Excel"""
    
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pendaftaran PPDB"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = [
            'No', 'Nomor Pendaftaran', 'Nama Lengkap', 'NIK', 'NISN',
            'Tempat Lahir', 'Tanggal Lahir', 'Jenis Kelamin', 'Agama',
            'Email', 'No. HP', 'Program', 'Status',
            'Nama Ayah', 'Pekerjaan Ayah', 'Nama Ibu', 'Pekerjaan Ibu',
            'Alamat', 'Kota', 'Provinsi',
            'Tanggal Daftar', 'Tanggal Submit', 'Tanggal Verifikasi'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        status_filter = request.GET.get('status')
        registrations = StudentRegistration.objects.all().order_by('-created_at')
        
        if status_filter:
            registrations = registrations.filter(status=status_filter)
        
        for idx, reg in enumerate(registrations, start=2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=reg.registration_number)
            ws.cell(row=idx, column=3, value=reg.full_name)
            ws.cell(row=idx, column=4, value=reg.nik)
            ws.cell(row=idx, column=5, value=reg.nisn or '-')
            ws.cell(row=idx, column=6, value=reg.birth_place)
            ws.cell(row=idx, column=7, value=reg.birth_date.strftime('%d/%m/%Y'))
            ws.cell(row=idx, column=8, value=reg.get_gender_display())
            ws.cell(row=idx, column=9, value=reg.get_religion_display())
            ws.cell(row=idx, column=10, value=reg.contact_email)
            ws.cell(row=idx, column=11, value=reg.contact_phone)
            ws.cell(row=idx, column=12, value=reg.get_program_choice_display())
            ws.cell(row=idx, column=13, value=reg.get_status_display())
            ws.cell(row=idx, column=14, value=reg.father_name)
            ws.cell(row=idx, column=15, value=reg.father_occupation)
            ws.cell(row=idx, column=16, value=reg.mother_name)
            ws.cell(row=idx, column=17, value=reg.mother_occupation)
            ws.cell(row=idx, column=18, value=reg.address)
            ws.cell(row=idx, column=19, value=reg.city)
            ws.cell(row=idx, column=20, value=reg.province)
            ws.cell(row=idx, column=21, value=reg.created_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=idx, column=22, value=reg.submitted_at.strftime('%d/%m/%Y %H:%M') if reg.submitted_at else '-')
            ws.cell(row=idx, column=23, value=reg.verified_at.strftime('%d/%m/%Y %H:%M') if reg.verified_at else '-')
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Pendaftaran_PPDB_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f"Export executed by {request.user}")
        
        return response
    
@require_POST
def delete_document_view(request, doc_id):
    """Hapus dokumen yang sudah diupload"""
    
    document = get_object_or_404(Document, pk=doc_id)
    registration = document.registration
    
    # Check status - hanya DRAFT yang bisa hapus dokumen
    if registration.status != StudentRegistration.RegistrationStatus.DRAFT:
        messages.error(request, 'Dokumen tidak bisa dihapus setelah pendaftaran disubmit.')
        return redirect('registration:documents', pk=registration.id)
    
    # Hapus file dari storage
    if document.file:
        document.file.delete()
    
    doc_type = document.get_document_type_display()
    document.delete()
    
    messages.success(request, f'Dokumen {doc_type} berhasil dihapus. Silakan upload ulang.')
    logger.info(f"Document deleted: {doc_type} for {registration.registration_number}")
    
    return redirect('registration:documents', pk=registration.id)
    
    # ============================================
# ERROR HANDLERS
# ============================================

def error_403(request, exception=None):
    """Custom 403 Forbidden page"""
    return render(request, 'errors/403.html', status=403)

def error_404(request, exception=None):
    """Custom 404 Not Found page"""
    return render(request, 'errors/404.html', status=404)

def error_500(request):
    """Custom 500 Server Error page"""
    return render(request, 'errors/500.html', status=500)

